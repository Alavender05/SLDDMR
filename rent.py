# rent.py
import openpyxl
import re
import math
import io
from openpyxl.utils import column_index_from_string


def parse_size_header(h):
    """Return numeric SQM size if header like '2.25 SQM', else None."""
    if not isinstance(h, str):
        return None
    h = h.strip()
    m = re.match(r"^(\d+(?:\.\d+)?)\s*SQM$", h, flags=re.IGNORECASE)
    return float(m.group(1)) if m else None


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s in {"", "-"}:
            return None
        s = s.replace(",", "")
        try:
            return float(s)
        except Exception:
            return None
    return None


def split_name_address(raw_name):
    """
    Raw competitor string is typically:
      'Brand - Site, 123 Street, Suburb, State Postcode'
    We split on the first comma.
    """
    if not isinstance(raw_name, str) or not raw_name.strip():
        return None, None
    parts = raw_name.split(",", 1)
    name = parts[0].strip()
    addr = parts[1].strip() if len(parts) > 1 else ""
    return name, addr


def extract_comps_from_raw(raw_file_bytes):
    """
    Reads the raw workbook and returns list of comps:
      {name, address, distance, values{size: val}}

    Assumptions (matches your current raw template):
      - Size headers row: 4, starting at column D
      - Competitor metric marker in column C: '12 mo. trailing avg.'
      - Name + distance stored in previous row (r-1) columns A and B
    """
    wb = openpyxl.load_workbook(raw_file_bytes, data_only=True)
    ws = wb.active

    # Map size -> column index
    size_col_map = {}
    for c in range(4, ws.max_column + 1):
        sz = parse_size_header(ws.cell(4, c).value)
        if sz is not None:
            size_col_map[sz] = c

    comps = []
    for r in range(5, ws.max_row + 1):
        stat = ws.cell(r, 3).value
        if isinstance(stat, str) and stat.strip().lower().startswith("12 mo. trailing avg"):
            raw_name = ws.cell(r - 1, 1).value
            dist = to_float(ws.cell(r - 1, 2).value)

            name, address = split_name_address(raw_name)

            values = {}
            for sz, c in size_col_map.items():
                v = to_float(ws.cell(r, c).value)
                if v is not None:
                    values[sz] = v

            if values:
                comps.append(
                    {
                        "name": name,
                        "address": address,
                        "distance": dist,
                        "values": values,
                    }
                )

    return comps


def fill_template(template_file_bytes, comps, max_comp_slots=16):
    """
    Writes into Template Market Rent Analysis 2:
      - Sheet: 'Comps & Unit Mix'
      - Competitor name across row starting N3 (N3, P3, R3, ...)
      - Competitor address across row starting N4 (N4, P4, R4, ...)
      - Distance into row 9 of asking-rate column (N9, P9, ...)
      - Mark selected Yes in row 10 of asking-rate column (N10, P10, ...)
      - Fill asking rates by matching SM size in column C (rows 12..)
      - Asking-rate columns are N, P, R... (every 2 columns)
    """
    wb = openpyxl.load_workbook(template_file_bytes)
    if "Comps & Unit Mix" not in wb.sheetnames:
        raise ValueError("Template workbook must contain a sheet named 'Comps & Unit Mix'.")

    ws = wb["Comps & Unit Mix"]

    # Map SM size (col C) -> row
    size_row_map = {}
    for r in range(12, 300):
        sm = to_float(ws.cell(r, 3).value)  # column C
        if sm is not None:
            size_row_map[sm] = r

    first_ask_col = column_index_from_string("N")

    for i, comp in enumerate(comps[:max_comp_slots]):
        ask_col = first_ask_col + 2 * i

        # Name and address rows
        ws.cell(3, ask_col).value = comp.get("name")
        ws.cell(4, ask_col).value = comp.get("address")

        # Distance row
        if comp.get("distance") is not None:
            ws.cell(9, ask_col).value = comp["distance"]

        # Selected flag
        ws.cell(10, ask_col).value = "Yes"

        # Fill rates by size
        for sz, val in comp["values"].items():
            row = size_row_map.get(sz)
            if row is not None:
                ws.cell(row, ask_col).value = val

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
    