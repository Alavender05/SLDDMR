import copy as _copy
import difflib
import hashlib
import io
import math
import re
import statistics
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.utils import column_index_from_string

_SIZE_RE = re.compile(r"^(\d+(?:\.\d+)?)\s*SQM$", re.IGNORECASE)


# ==========================================
# CORE LOGIC
# ==========================================

def parse_size_header(h):
    """Return numeric SQM size if header like '2.25 SQM', else None."""
    if not isinstance(h, str):
        return None
    m = _SIZE_RE.match(h.strip())
    return float(m.group(1)) if m else None


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s in {"", "-"}:
            return None
        s = s.replace(",", "")
        try:
            return float(s)
        except Exception:
            return None
    return None


def split_name_address(raw_name):
    """
    Raw competitor string is typically:
      'Brand - Site, 123 Street, Suburb, State Postcode'
    We split on the first comma.
    """
    if not isinstance(raw_name, str) or not raw_name.strip():
        return None, None
    parts = raw_name.split(",", 1)
    name = parts[0].strip()
    addr = parts[1].strip() if len(parts) > 1 else ""
    return name, addr


def find_data_sheet(wb):
    """
    Return the worksheet that contains StoreTrack data by scanning all sheets
    for '12 mo. trailing avg.' in column C. Falls back to the active sheet.
    """
    for ws in wb.worksheets:
        for r in range(5, min(ws.max_row + 1, 200)):
            v = ws.cell(r, 3).value
            if isinstance(v, str) and v.strip().lower().startswith("12 mo. trailing avg"):
                return ws
    return wb.active


def extract_comps_from_raw(raw_file_bytes):
    """
    Reads the raw workbook and returns list of comps:
      {name, address, distance, values{size: val}}

    Assumptions:
      - Size headers row: 4, starting at column D
      - Competitor metric marker in column C: '12 mo. trailing avg.'
      - Name + distance stored in previous row (r-1) columns A and B
    """
    wb = openpyxl.load_workbook(raw_file_bytes, data_only=True)
    ws = find_data_sheet(wb)

    size_col_map = {}
    for c in range(4, ws.max_column + 1):
        sz = parse_size_header(ws.cell(4, c).value)
        if sz is not None:
            size_col_map[sz] = c

    comps = []
    for r in range(5, ws.max_row + 1):
        stat = ws.cell(r, 3).value
        if isinstance(stat, str) and stat.strip().lower().startswith("12 mo. trailing avg"):
            raw_name = ws.cell(r - 1, 1).value
            dist = to_float(ws.cell(r - 1, 2).value)
            name, address = split_name_address(raw_name)

            values = {}
            for sz, c in size_col_map.items():
                v = to_float(ws.cell(r, c).value)
                if v is not None:
                    values[sz] = v

            if values:
                comps.append({
                    "name": name,
                    "address": address,
                    "distance": dist,
                    "values": values,
                })

    return comps


def build_level_row_maps(ws):
    """
    Builds row maps for each floor level by reading the Type column (E)
    to classify rows into Ground, Upper, Drive-up, and Vehicle/RV sections.

    The template has multiple contiguous size blocks separated by header
    rows (where column C = 'SM'). Each data row has:
      - Column C: SQM value (float)
      - Column D: Floor number (1 = ground level, 2 = upper level)
      - Column E: Type string ('Drive-up', 'Ground', 'Upper', 'Vehicle', 'RV', 'Container')

    Ground Floor StoreTrack data maps to Type='Ground' rows.
    Upper Level StoreTrack data maps to Type='Upper' rows.

    Returns:
        gf_map       - {sqm_float: row_int} for Ground rows
        ul_map       - {sqm_float: row_int} for Upper rows
        dup_warnings - list of strings for duplicate SQM sizes within a section
    """
    gf_map: dict[float, int] = {}
    ul_map: dict[float, int] = {}
    gf_dups: list[str] = []
    ul_dups: list[str] = []

    for r in range(12, ws.max_row + 1):
        sm = to_float(ws.cell(r, 3).value)  # Column C: SQM
        if sm is None:
            continue

        type_val = ws.cell(r, 5).value  # Column E: Type
        if not isinstance(type_val, str):
            continue
        type_val = type_val.strip().lower()

        if type_val == "ground":
            if sm in gf_map:
                gf_dups.append(
                    f"{sm} SQM appears more than once in the Ground section "
                    f"(rows {gf_map[sm]} and {r}); first occurrence used."
                )
            else:
                gf_map[sm] = r
        elif type_val == "upper":
            if sm in ul_map:
                ul_dups.append(
                    f"{sm} SQM appears more than once in the Upper section "
                    f"(rows {ul_map[sm]} and {r}); first occurrence used."
                )
            else:
                ul_map[sm] = r
        # Drive-up, Vehicle, RV, Container rows are skipped for now
        # (not populated from StoreTrack GF/UL exports)

    return gf_map, ul_map, gf_dups + ul_dups


def fuzzy_match_name(name: str, candidates: dict, threshold: float = 0.85):
    """
    Find the best fuzzy match for `name` among the (normalised) keys of
    `candidates` using difflib.SequenceMatcher.

    Returns (best_key, score) if score >= threshold, else (None, 0.0).
    """
    best_key, best_score = None, 0.0
    for k in candidates:
        score = difflib.SequenceMatcher(None, name, k).ratio()
        if score > best_score:
            best_score = score
            best_key = k
    if best_score >= threshold:
        return best_key, best_score
    return None, 0.0


def align_comps(gf_comps, ul_comps):
    """
    Align GF and UL comp lists by competitor name using a two-pass strategy:
      1. Exact match after normalisation (strip + lowercase).
      2. Fuzzy match via SequenceMatcher (threshold 0.85) for unmatched names
         — generates a verification warning per fuzzy pair.

    Returns:
        slots    – list of {"gf": comp_or_None, "ul": comp_or_None}
        warnings – fuzzy-match and one-sided mismatch warning strings
    """
    def norm(name):
        return (name or "").strip().lower()

    gf_by_name = {norm(c["name"]): c for c in gf_comps}
    ul_by_name = {norm(c["name"]): c for c in ul_comps}

    # Pass 1: exact matches
    gf_to_ul: dict[str, str] = {}
    for k in gf_by_name:
        if k in ul_by_name:
            gf_to_ul[k] = k

    unmatched_gf = [k for k in gf_by_name if k not in gf_to_ul]
    unmatched_ul = set(ul_by_name) - set(gf_to_ul.values())

    # Pass 2: fuzzy matches for remaining unmatched names
    fuzzy_warnings: list[str] = []
    for gf_k in unmatched_gf:
        best_ul_k, score = fuzzy_match_name(gf_k, {k: None for k in unmatched_ul})
        if best_ul_k is not None:
            gf_to_ul[gf_k] = best_ul_k
            unmatched_ul.discard(best_ul_k)
            gf_display = gf_by_name[gf_k]["name"] or gf_k
            ul_display = ul_by_name[best_ul_k]["name"] or best_ul_k
            fuzzy_warnings.append(
                f"Fuzzy match: '{gf_display}' (GF) matched to '{ul_display}' (UL) "
                f"[{score:.0%} similarity] — verify these are the same facility."
            )

    # Build ordered slot list: GF order first, then UL-only extras
    ordered: list[tuple] = []
    seen: set[str] = set()
    for gf_k in gf_by_name:
        if gf_k not in seen:
            seen.add(gf_k)
            ordered.append((gf_k, gf_to_ul.get(gf_k)))

    matched_ul = set(gf_to_ul.values())
    for ul_k in ul_by_name:
        if ul_k not in matched_ul and ul_k not in seen:
            seen.add(ul_k)
            ordered.append((None, ul_k))

    slots: list[dict] = []
    mismatch_warnings: list[str] = []
    for gf_k, ul_k in ordered:
        gf = gf_by_name.get(gf_k) if gf_k else None
        ul = ul_by_name.get(ul_k) if ul_k else None
        display = (gf or ul)["name"] or (gf_k or ul_k)  # type: ignore[index]
        if gf and not ul and ul_comps:
            mismatch_warnings.append(
                f"'{display}' found in Ground Floor data only — Upper Level rates will be blank."
            )
        elif ul and not gf and gf_comps:
            mismatch_warnings.append(
                f"'{display}' found in Upper Level data only — Ground Floor rates will be blank."
            )
        slots.append({"gf": gf, "ul": ul})

    return slots, fuzzy_warnings + mismatch_warnings


def validate_floor_assignment(gf_comps, ul_comps):
    """
    Heuristic checks that GF and UL files were not uploaded to the wrong slots.

    Heuristic 1: GF mean rate is more than 10% below UL mean rate.
                 Ground Floor storage is conventionally priced higher.
    Heuristic 2: GF and UL rates are near-identical across ≥80% of
                 size/competitor intersections (same file uploaded twice).

    Returns a list of warning strings (empty if no issues detected).
    """
    warnings: list[str] = []
    if not gf_comps or not ul_comps:
        return warnings

    gf_vals = [v for c in gf_comps for v in c["values"].values()]
    ul_vals = [v for c in ul_comps for v in c["values"].values()]

    if gf_vals and ul_vals:
        gf_mean = statistics.mean(gf_vals)
        ul_mean = statistics.mean(ul_vals)
        if ul_mean > 0 and gf_mean < ul_mean * 0.90:
            warnings.append(
                f"Ground Floor mean rate (${gf_mean:.2f}) is more than 10% below "
                f"Upper Level (${ul_mean:.2f}). Ground Floor storage is typically "
                "priced higher — verify that files are in the correct uploaders."
            )

    gf_by_name = {(c["name"] or "").strip().lower(): c for c in gf_comps}
    ul_by_name = {(c["name"] or "").strip().lower(): c for c in ul_comps}
    matches = comparisons = 0
    for k in gf_by_name:
        if k in ul_by_name:
            for sz, gv in gf_by_name[k]["values"].items():
                uv = ul_by_name[k]["values"].get(sz)
                if uv is not None:
                    comparisons += 1
                    if abs(gv - uv) < 0.01:
                        matches += 1
    if comparisons >= 3 and matches / comparisons >= 0.80:
        warnings.append(
            f"Ground Floor and Upper Level rates are identical for {matches}/{comparisons} "
            "size/competitor intersections — the same file may have been uploaded to both inputs."
        )

    return warnings


def get_max_comp_slots(ws):
    """Derive how many competitor column slots fit in the sheet from column N onward."""
    return len(find_asking_rate_columns(ws)) or 16


def find_asking_rate_columns(ws):
    """
    Return a list of column indices that have 'Asking Rate' in row 11.
    These are the competitor data columns (N, P, R, T, ...).

    Scans up to 200 columns from N to avoid iterating through the full
    16,383 Excel column range when the sheet has far-right formatting.
    Stops early once 20 consecutive empty columns are encountered.
    """
    first_ask_col = column_index_from_string("N")
    max_scan = min(first_ask_col + 200, ws.max_column + 1)
    cols = []
    empty_streak = 0
    for c in range(first_ask_col, max_scan):
        v = ws.cell(11, c).value
        if isinstance(v, str) and v.strip() == "Asking Rate":
            cols.append(c)
            empty_streak = 0
        elif v is None or (isinstance(v, str) and not v.strip()):
            empty_streak += 1
            if empty_streak > 20:
                break
        else:
            empty_streak = 0
    return cols


def find_closest_size(sz: float, size_map: dict, tolerance: float = 0.05):
    """
    Find the template row and matched key for the closest SQM value within
    `tolerance`. Returns (row_int, matched_sz) on success, (None, None) if
    nothing is within tolerance.
    """
    if sz in size_map:
        return size_map[sz], sz
    best_key, best_diff = None, float("inf")
    for k in size_map:
        diff = abs(k - sz)
        if diff < best_diff:
            best_diff = diff
            best_key = k
    if best_key is not None and best_diff <= tolerance:
        return size_map[best_key], best_key
    return None, None


def _copy_cell_style(source_cell, target_cell):
    """Copy font, fill, alignment, and border from source to target cell."""
    if source_cell.has_style:
        target_cell.font = _copy.copy(source_cell.font)
        target_cell.fill = _copy.copy(source_cell.fill)
        target_cell.alignment = _copy.copy(source_cell.alignment)
        target_cell.border = _copy.copy(source_cell.border)


def fill_template(template_file_bytes, slots, max_comp_slots=None):
    """
    Writes competitor rates into the template using pre-aligned slots.

    Each slot is {"gf": comp_or_None, "ul": comp_or_None}.

    Sheet: 'Comps & Unit Mix'
      - Competitor name  → row 3,  Asking Rate columns (N, P, R, ...)
      - Competitor addr  → row 4,  same columns
      - Distance         → row 9,  same columns
      - Selected flag    → row 10, same columns  ('Yes')
      - Ground rates     → Type='Ground' rows in col C (identified via col E)
      - Upper rates      → Type='Upper' rows in col C (identified via col E)

    Rate cells receive currency formatting ($#,##0.00).
    Size matching uses a ±0.05 SQM tolerance to handle minor float differences.

    Returns:
        out       – BytesIO of the filled workbook
        n_written – number of competitor slots written
        warnings  – combined duplicate-SQM and size-match warning strings
    """
    wb = openpyxl.load_workbook(template_file_bytes)
    if "Comps & Unit Mix" not in wb.sheetnames:
        raise ValueError("Template workbook must contain a sheet named 'Comps & Unit Mix'.")

    ws = wb["Comps & Unit Mix"]
    gf_map, ul_map, dup_warnings = build_level_row_maps(ws)

    # Get the actual Asking Rate columns from row 11
    ask_columns = find_asking_rate_columns(ws)
    if not ask_columns:
        # Fallback to old N, P, R pattern
        first_ask_col = column_index_from_string("N")
        ask_columns = [first_ask_col + 2 * i for i in range(16)]

    if max_comp_slots is None:
        max_comp_slots = len(ask_columns)

    n_written = 0
    size_match_warnings: list[str] = []

    for i, slot in enumerate(slots[:max_comp_slots]):
        if i >= len(ask_columns):
            break
        ask_col = ask_columns[i]
        gf_comp = slot["gf"]
        ul_comp = slot["ul"]
        identity = gf_comp or ul_comp
        comp_name = identity.get("name") or f"Competitor {i + 1}"

        # Write competitor header info
        ws.cell(3, ask_col).value = identity.get("name")
        ws.cell(4, ask_col).value = identity.get("address")
        if identity.get("distance") is not None:
            ws.cell(9, ask_col).value = identity["distance"]
        ws.cell(10, ask_col).value = "Yes"

        if gf_comp:
            for sz, val in gf_comp["values"].items():
                row, matched_sz = find_closest_size(sz, gf_map)
                if row is not None:
                    ws.cell(row, ask_col).value = val
                    ws.cell(row, ask_col).number_format = '"$"#,##0.00'
                    if matched_sz is not None and abs(matched_sz - sz) > 1e-9:
                        size_match_warnings.append(
                            f"GF — {comp_name}: {sz} SQM matched to {matched_sz} SQM template row."
                        )

        if ul_comp:
            for sz, val in ul_comp["values"].items():
                row, matched_sz = find_closest_size(sz, ul_map)
                if row is not None:
                    ws.cell(row, ask_col).value = val
                    ws.cell(row, ask_col).number_format = '"$"#,##0.00'
                    if matched_sz is not None and abs(matched_sz - sz) > 1e-9:
                        size_match_warnings.append(
                            f"UL — {comp_name}: {sz} SQM matched to {matched_sz} SQM template row."
                        )

        n_written += 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, n_written, dup_warnings + size_match_warnings


# ==========================================
# STREAMLIT UI HELPERS
# ==========================================

def _file_signature(f) -> str | None:
    """Stable fingerprint for an uploaded file (name + size). Returns None if no file."""
    if f is None:
        return None
    return hashlib.md5((f.name + str(f.size)).encode()).hexdigest()


def _render_comp_preview(comps: list, label: str):
    """Render competitor data as a dataframe in a collapsible expander."""
    if not comps:
        return
    all_sizes = sorted({sz for c in comps for sz in c["values"]})
    rows = []
    for c in comps:
        row: dict = {
            "Name": c["name"] or "Unknown",
            "Address": c["address"] or "—",
            "Distance (km)": c["distance"] if c["distance"] is not None else "N/A",
        }
        for sz in all_sizes:
            val = c["values"].get(sz)
            row[f"{sz} SQM"] = f"${val:.2f}" if val is not None else "—"
        rows.append(row)
    with st.expander(f"Preview — {label} ({len(comps)} competitor(s))"):
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


def _render_slot_editor(slots: list) -> list:
    """
    Let the analyst include/exclude and reorder competitors before generating.
    Renders a st.data_editor with Include (checkbox) and Order (number) editable.
    Returns the filtered and reordered slot list.
    """
    rows = []
    for i, slot in enumerate(slots, 1):
        identity = slot["gf"] or slot["ul"]
        rows.append({
            "Include": True,
            "Order": i,
            "Name": identity.get("name") or "Unknown",
            "Distance (km)": (
                identity["distance"] if identity.get("distance") is not None else "N/A"
            ),
            "GF Sizes (SQM)": (
                ", ".join(str(sz) for sz in sorted(slot["gf"]["values"]))
                if slot["gf"] else "—"
            ),
            "UL Sizes (SQM)": (
                ", ".join(str(sz) for sz in sorted(slot["ul"]["values"]))
                if slot["ul"] else "—"
            ),
        })

    df = pd.DataFrame(rows)
    with st.expander("Select and Order Competitors", expanded=True):
        st.caption(
            "Tick/untick **Include** to add or remove competitors. "
            "Edit **Order** to change the column position in the template."
        )
        edited = st.data_editor(
            df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Include": st.column_config.CheckboxColumn("Include", default=True),
                "Order": st.column_config.NumberColumn("Order", min_value=1, step=1),
                "Name": st.column_config.TextColumn("Name", disabled=True),
                "Distance (km)": st.column_config.TextColumn("Distance (km)", disabled=True),
                "GF Sizes (SQM)": st.column_config.TextColumn("GF Sizes (SQM)", disabled=True),
                "UL Sizes (SQM)": st.column_config.TextColumn("UL Sizes (SQM)", disabled=True),
            },
            num_rows="fixed",
        )

    included = edited[edited["Include"]].sort_values("Order")
    return [slots[i] for i in included.index.tolist()]


def _compute_linear_regression(x_vals, y_vals):
    """
    OLS linear regression using stdlib only.
    Returns (slope, intercept, r, r_squared) or None if fewer than 2 points
    or zero variance in x.
    """
    n = len(x_vals)
    if n < 2:
        return None
    x_mean = statistics.mean(x_vals)
    y_mean = statistics.mean(y_vals)
    ss_xy = sum((x - x_mean) * (y - y_mean) for x, y in zip(x_vals, y_vals))
    ss_xx = sum((x - x_mean) ** 2 for x in x_vals)
    ss_yy = sum((y - y_mean) ** 2 for y in y_vals)
    if ss_xx == 0:
        return None
    slope = ss_xy / ss_xx
    intercept = y_mean - slope * x_mean
    r = ss_xy / math.sqrt(ss_xx * ss_yy) if ss_yy > 0 else 0.0
    return slope, intercept, r, r ** 2


def _build_regression_figure(data: dict, title: str, subject_data: dict | None = None):
    """
    Build a matplotlib scatter + regression line figure.

    data: {sqm_float: [price_float, ...]} — one price per competitor per size.
    Each (sqm, price) pair becomes one scatter point, coloured by size bucket.
    subject_data: {sqm_float: price_float} — subject property rates; plotted as gold stars.
    The regression is fitted across all competitor points only.
    """
    x_all, y_all = [], []
    for sz in sorted(data):
        for price in data[sz]:
            x_all.append(sz)
            y_all.append(price)

    fig, ax = plt.subplots(figsize=(7, 5))

    unique_sizes = sorted(data)
    cmap = plt.colormaps.get_cmap("tab10")
    for idx, sz in enumerate(unique_sizes):
        ax.scatter(
            [sz] * len(data[sz]), data[sz],
            color=cmap(idx % 10), s=70, zorder=3,
            label=f"{sz} SQM",
        )

    if subject_data:
        sx = sorted(subject_data)
        sy = [subject_data[sz] for sz in sx]
        ax.scatter(
            sx, sy, color="gold", s=220, zorder=6, marker="*",
            edgecolors="black", linewidths=0.7, label="Subject Property",
        )

    reg = _compute_linear_regression(x_all, y_all)
    if reg:
        slope, intercept, r, r_squared = reg
        x_min, x_max = min(x_all), max(x_all)
        ys = [slope * x + intercept for x in (x_min, x_max)]
        ax.plot([x_min, x_max], ys, color="crimson", linewidth=2, zorder=4, label="Regression")

        sign = "+" if intercept >= 0 else "-"
        eq_text = (
            f"y = {slope:.2f}x {sign} {abs(intercept):.2f}\n"
            f"R² = {r_squared:.4f}\n"
            f"r  = {r:.4f}"
        )
        ax.text(
            0.97, 0.05, eq_text,
            transform=ax.transAxes,
            fontsize=9, verticalalignment="bottom", horizontalalignment="right",
            bbox=dict(boxstyle="round,pad=0.4", facecolor="white", alpha=0.85, edgecolor="#aaaaaa"),
        )

    ax.set_xlabel("Size (SQM)", fontsize=11)
    ax.set_ylabel("Price ($)", fontsize=11)
    ax.set_title(title, fontsize=12, fontweight="bold")
    ax.legend(fontsize=8, loc="upper right")
    ax.grid(True, linestyle="--", alpha=0.4)
    plt.tight_layout()
    return fig


def _render_market_summary(slots: list, gf_map: dict, ul_map: dict,
                           subject_gf: dict | None = None, subject_ul: dict | None = None):
    """
    Display min/max/mean/median/Q1/Q3 per SQM size across all competitors, flag
    Tukey IQR outliers, show regression charts with subject property overlay, and
    render a colour-coded rate heatmap (competitor × SQM).
    """
    st.subheader("Market Rate Summary")

    def gather_vals(size_map: dict, level_key: str) -> dict[float, list[float]]:
        data: dict[float, list[float]] = {}
        for sz in size_map:
            for slot in slots:
                comp = slot.get(level_key)
                if not comp:
                    continue
                v = comp["values"].get(sz)
                if v is None:
                    for k, kv in comp["values"].items():
                        if abs(k - sz) <= 0.05:
                            v = kv
                            break
                if v is not None:
                    data.setdefault(sz, []).append(v)
        return data

    def build_stats_df(data: dict[float, list[float]]) -> tuple[pd.DataFrame, list[str]]:
        rows, outlier_msgs = [], []
        for sz in sorted(data):
            vals = data[sz]
            avg = statistics.mean(vals)
            med = statistics.median(vals)
            row = {
                "Size (SQM)": sz,
                "Count": len(vals),
                "Min ($)": f"${min(vals):.2f}",
                "Max ($)": f"${max(vals):.2f}",
                "Mean ($)": f"${avg:.2f}",
                "Median ($)": f"${med:.2f}",
                "Q1 ($)": "—",
                "Q3 ($)": "—",
            }
            if len(vals) >= 4:
                qs = statistics.quantiles(vals, n=4)
                q1, q3 = qs[0], qs[2]
                iqr_val = q3 - q1
                fence_lo = q1 - 1.5 * iqr_val
                fence_hi = q3 + 1.5 * iqr_val
                row["Q1 ($)"] = f"${q1:.2f}"
                row["Q3 ($)"] = f"${q3:.2f}"
                for v in vals:
                    if v < fence_lo or v > fence_hi:
                        outlier_msgs.append(
                            f"{sz} SQM: ${v:.2f} is outside Tukey IQR fence "
                            f"(${fence_lo:.2f} – ${fence_hi:.2f})"
                        )
            elif len(vals) >= 3:
                try:
                    sd = statistics.stdev(vals)
                    for v in vals:
                        if abs(v - avg) > 2 * sd:
                            outlier_msgs.append(
                                f"{sz} SQM: ${v:.2f} is >2 SD from mean "
                                f"(${avg:.2f} ± ${sd:.2f}) [small sample]"
                            )
                except statistics.StatisticsError:
                    pass
            rows.append(row)
        return pd.DataFrame(rows), outlier_msgs

    gf_data = gather_vals(gf_map, "gf")
    ul_data = gather_vals(ul_map, "ul")
    all_outliers: list[str] = []

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Ground Floor**")
        if gf_data:
            df, msgs = build_stats_df(gf_data)
            st.dataframe(df, use_container_width=True, hide_index=True)
            all_outliers.extend(f"GF — {m}" for m in msgs)
            fig = _build_regression_figure(gf_data, "Ground Floor — Rent vs Size", subject_gf)
            st.pyplot(fig, use_container_width=True)
            plt.close(fig)
        else:
            st.info("No Ground Floor rate data available.")

    with col2:
        st.markdown("**Upper Level**")
        if ul_data:
            df, msgs = build_stats_df(ul_data)
            st.dataframe(df, use_container_width=True, hide_index=True)
            all_outliers.extend(f"UL — {m}" for m in msgs)
            fig = _build_regression_figure(ul_data, "Upper Level — Rent vs Size", subject_ul)
            st.pyplot(fig, use_container_width=True)
            plt.close(fig)
        else:
            st.info("No Upper Level rate data available.")

    for msg in all_outliers:
        st.warning(f"Outlier detected: {msg}")

    # ---- Rate Heatmap ----
    st.markdown("#### Competitor Rate Heatmap")
    st.caption(
        "Colour scale per column: green = lower rate, red = higher rate. "
        "★ Subject Property row is highlighted in amber."
    )

    def build_pivot(level_key: str, size_map: dict, subject_data: dict | None) -> pd.DataFrame:
        sizes = sorted(size_map)
        col_names = [f"{sz:g} SQM" for sz in sizes]
        rows = []
        for slot in slots:
            comp = slot.get(level_key)
            if not comp:
                continue
            row = {"Competitor": comp["name"] or "Unknown"}
            for sz, col in zip(sizes, col_names):
                v = comp["values"].get(sz)
                if v is None:
                    for k, kv in comp["values"].items():
                        if abs(k - sz) <= 0.05:
                            v = kv
                            break
                row[col] = v
            rows.append(row)
        if subject_data:
            subject_row = {"Competitor": "★ Subject Property"}
            for sz, col in zip(sizes, col_names):
                subject_row[col] = subject_data.get(sz)
            rows.append(subject_row)
        return pd.DataFrame(rows)

    def render_heatmap(level_key: str, size_map: dict, subject_data: dict | None):
        df = build_pivot(level_key, size_map, subject_data)
        if df.empty:
            st.info("No data.")
            return
        size_cols = [c for c in df.columns if c != "Competitor"]
        display = df.set_index("Competitor")

        def highlight_subject(row):
            if row.name == "★ Subject Property":
                return ["background-color: #fff8e1; font-weight: bold"] * len(row)
            return [""] * len(row)

        comp_mask = display.index != "★ Subject Property"
        styled = (
            display[size_cols]
            .style
            .background_gradient(cmap="RdYlGn_r", axis=0,
                                 subset=pd.IndexSlice[display.index[comp_mask], :])
            .apply(highlight_subject, axis=1)
            .format(lambda x: f"${x:.2f}" if pd.notna(x) else "—")
        )
        st.dataframe(styled, use_container_width=True)

    tab_gf, tab_ul = st.tabs(["Ground Floor", "Upper Level"])
    with tab_gf:
        render_heatmap("gf", gf_map, subject_gf)
    with tab_ul:
        render_heatmap("ul", ul_map, subject_ul)


# ==========================================
# STREAMLIT APP
# ==========================================

def main():
    st.set_page_config(
        page_title="Market Rent Transposer",
        layout="wide",
        page_icon="🏢",
    )
    st.title("Market Rent Transposer")
    st.write(
        "Upload the Ground Floor and Upper Level StoreTrack exports plus the "
        "Market Rent Analysis template. The app will extract competitor rates "
        "from each file and insert them into the matching floor-level rows of "
        "the template."
    )

    col1, col2, col3 = st.columns(3)

    with col1:
        st.subheader("1. Ground Floor Raw Data")
        gf_file = st.file_uploader(
            "Upload Ground Floor StoreTrack export (.xlsx)",
            type=["xlsx"],
            key="gf",
            help="Raw workbook with size headers in row 4 (column D onward) "
                 "and '12 mo. trailing avg.' rows per competitor — Ground Floor pricing.",
        )

    with col2:
        st.subheader("2. Upper Level Raw Data")
        ul_file = st.file_uploader(
            "Upload Upper Level StoreTrack export (.xlsx)",
            type=["xlsx"],
            key="ul",
            help="Same format as the Ground Floor file — Upper Level pricing.",
        )

    with col3:
        st.subheader("3. Template")
        template_file = st.file_uploader(
            "Upload Market Rent Analysis template (.xlsx)",
            type=["xlsx"],
            key="tmpl",
            help="Must contain a sheet named 'Comps & Unit Mix'. "
                 "Ground rows identified by Type='Ground' in column E; "
                 "Upper rows by Type='Upper'.",
        )

    st.markdown("---")

    if not template_file:
        st.info("Please upload the template file to continue.")
        return
    if not gf_file and not ul_file:
        st.info("Please upload at least one raw data file (Ground Floor or Upper Level).")
        return

    # ---- Stale session state invalidation ----
    current_sig = (
        _file_signature(gf_file),
        _file_signature(ul_file),
        _file_signature(template_file),
    )
    if st.session_state.get("file_sig") != current_sig:
        for key in ("result", "n_written", "dup_warnings"):
            st.session_state.pop(key, None)
        st.session_state["file_sig"] = current_sig

    # ---- Wrap uploads in BytesIO immediately ----
    gf_bytes = io.BytesIO(gf_file.read()) if gf_file else None
    ul_bytes = io.BytesIO(ul_file.read()) if ul_file else None
    tmpl_bytes = io.BytesIO(template_file.read())

    # ---- Extract comps ----
    gf_comps: list = []
    ul_comps: list = []

    if gf_bytes:
        with st.spinner("Reading Ground Floor data..."):
            try:
                gf_comps = extract_comps_from_raw(gf_bytes)
            except Exception as e:
                st.error(f"Could not read Ground Floor file: {e}")
                return
        if not gf_comps:
            st.warning(
                "No competitors found in the Ground Floor file. "
                "Check that row 4 has SQM headers and column C has '12 mo. trailing avg.' rows."
            )
        else:
            st.success(f"Ground Floor: found **{len(gf_comps)}** competitor(s).")
            _render_comp_preview(gf_comps, "Ground Floor")

    if ul_bytes:
        with st.spinner("Reading Upper Level data..."):
            try:
                ul_comps = extract_comps_from_raw(ul_bytes)
            except Exception as e:
                st.error(f"Could not read Upper Level file: {e}")
                return
        if not ul_comps:
            st.warning(
                "No competitors found in the Upper Level file. "
                "Check that row 4 has SQM headers and column C has '12 mo. trailing avg.' rows."
            )
        else:
            st.success(f"Upper Level: found **{len(ul_comps)}** competitor(s).")
            _render_comp_preview(ul_comps, "Upper Level")

    if not gf_comps and not ul_comps:
        st.error("No competitor data could be extracted from either file.")
        return

    # ---- Align competitors by name (exact + fuzzy) ----
    slots, align_warnings = align_comps(gf_comps, ul_comps)
    for w in align_warnings:
        st.warning(w)

    # ---- Floor assignment validation ----
    for w in validate_floor_assignment(gf_comps, ul_comps):
        st.warning(w)

    # ---- Peek at template: extract row maps and max slots ----
    gf_map: dict = {}
    ul_map: dict = {}
    subject_gf: dict = {}
    subject_ul: dict = {}
    max_slots = 16
    try:
        tmpl_bytes.seek(0)
        wb_peek = openpyxl.load_workbook(tmpl_bytes, data_only=True)
        if "Comps & Unit Mix" in wb_peek.sheetnames:
            ws_peek = wb_peek["Comps & Unit Mix"]
            max_slots = get_max_comp_slots(ws_peek)
            gf_map, ul_map, _ = build_level_row_maps(ws_peek)
            # Subject property rates: check columns G (Per SM under SELECTED COMPS)
            # In the new template, column G holds the subject's selected per-SM rate
            subject_col = column_index_from_string("G")
            for sz, r in gf_map.items():
                v = to_float(ws_peek.cell(r, subject_col).value)
                if v is not None:
                    subject_gf[sz] = v
            for sz, r in ul_map.items():
                v = to_float(ws_peek.cell(r, subject_col).value)
                if v is not None:
                    subject_ul[sz] = v
        tmpl_bytes.seek(0)
    except Exception:
        tmpl_bytes.seek(0)

    st.markdown("---")

    # ---- Competitor selection & reordering ----
    slots = _render_slot_editor(slots)

    if not slots:
        st.warning("No competitors selected. Tick at least one competitor to generate the report.")
        return

    if len(slots) > max_slots:
        st.warning(
            f"{max_slots} of {len(slots)} competitors will be written — "
            f"the template only has {max_slots} competitor column slot(s). "
            "Deselect extras in the table above."
        )

    # ---- Market rate summary + outlier detection ----
    if gf_map or ul_map:
        st.markdown("---")
        _render_market_summary(slots, gf_map, ul_map, subject_gf, subject_ul)

    # ---- Generate ----
    st.markdown("---")
    if st.button("Generate Filled Template", type="primary"):
        with st.spinner("Filling template..."):
            try:
                result, n_written, all_warnings = fill_template(tmpl_bytes, slots)
            except ValueError as e:
                st.error(str(e))
                return
            except Exception as e:
                st.error(f"Error filling template: {e}")
                return

        st.session_state["result"] = result
        st.session_state["n_written"] = n_written
        st.session_state["dup_warnings"] = all_warnings

    # Render download button outside the generate block so that clicking it
    # does not re-trigger generation on the script rerun Streamlit performs.
    if "result" in st.session_state:
        for w in st.session_state["dup_warnings"]:
            st.warning(w)
        st.success(f"Template filled with **{st.session_state['n_written']}** competitor(s).")
        st.download_button(
            label="Download Filled Template",
            data=st.session_state["result"],
            file_name="Market_Rent_Analysis_Filled.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()